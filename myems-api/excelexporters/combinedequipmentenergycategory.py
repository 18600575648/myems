import base64
import gettext
import os
import re
import uuid

import openpyxl.utils.cell as format_cell
from openpyxl import Workbook
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excel file
# Step 3: Encode the excel file to Base64
########################################################################################################################


def export(report,
           name,
           reporting_start_datetime_local,
           reporting_end_datetime_local,
           period_type,
           language):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if report is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(report,
                              name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local,
                              period_type,
                              language)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report,
                   name,
                   reporting_start_datetime_local,
                   reporting_end_datetime_local,
                   period_type,
                   language):

    locale_path = './i18n/'
    if language == 'zh_CN':
        trans = gettext.translation('myems', locale_path, languages=['zh_CN'])
    elif language == 'de':
        trans = gettext.translation('myems', locale_path, languages=['de'])
    elif language == 'en':
        trans = gettext.translation('myems', locale_path, languages=['en'])
    else:
        trans = gettext.translation('myems', locale_path, languages=['en'])
    trans.install()
    _ = trans.gettext

    wb = Workbook()
    ws = wb.active
    ws.title = "CombinedEquipmentEnergyCategory"

    # Row height
    ws.row_dimensions[1].height = 102

    for i in range(2, 2000 + 1):
        ws.row_dimensions[i].height = 42

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='90ee90')
    f_border = Border(left=Side(border_style='medium'),
                      right=Side(border_style='medium'),
                      bottom=Side(border_style='medium'),
                      top=Side(border_style='medium')
                      )
    b_border = Border(
        bottom=Side(border_style='medium'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = _('Name') + ':'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = name

    ws['D3'].alignment = b_r_alignment
    ws['D3'] = _('Period Type') + ':'
    ws['E3'].border = b_border
    ws['E3'].alignment = b_c_alignment
    ws['E3'] = period_type

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = _('Reporting Start Datetime') + ':'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['D4'].alignment = b_r_alignment
    ws['D4'] = _('Reporting End Datetime') + ':'
    ws['E4'].border = b_border
    ws['E4'].alignment = b_c_alignment
    ws['E4'] = reporting_end_datetime_local

    if "reporting_period" not in report.keys() or \
            "names" not in report['reporting_period'].keys() or len(report['reporting_period']['names']) == 0:
        filename = str(uuid.uuid4()) + '.xlsx'
        wb.save(filename)

        return filename
    ####################################################################################################################
    # First: Consumption
    # 6: title
    # 7: table title
    # 8~10 table_data
    # Total: 5 rows
    # if has not energy data: set low height for rows
    ####################################################################################################################
    report['reporting_period'] = report['reporting_period']
    if "names" not in report['reporting_period'].keys() or \
            report['reporting_period']['names'] is None or \
            len(report['reporting_period']['names']) == 0:
        for i in range(6, 9 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B6'].font = title_font
        ws['B6'] = name + ' ' + _('Consumption')

        category = report['reporting_period']['names']
        ca_len = len(category)

        ws.row_dimensions[7].height = 60
        ws['B7'].fill = table_fill
        ws['B7'].border = f_border

        ws['B8'].font = title_font
        ws['B8'].alignment = c_c_alignment
        ws['B8'] = _('Consumption')
        ws['B8'].border = f_border

        ws['B9'].font = title_font
        ws['B9'].alignment = c_c_alignment
        ws['B9'] = _('Increment Rate')
        ws['B9'].border = f_border

        col = ''

        for i in range(0, ca_len):
            col = chr(ord('C') + i)
            row = '7'
            cell = col + row
            ws[col + '7'].fill = table_fill
            ws[col + '7'].font = name_font
            ws[col + '7'].alignment = c_c_alignment
            ws[col + '7'] = report['reporting_period']['names'][i] + " (" + report['reporting_period']['units'][i] + ")"
            ws[col + '7'].border = f_border

            ws[col + '8'].font = name_font
            ws[col + '8'].alignment = c_c_alignment
            ws[col + '8'] = round(report['reporting_period']['subtotals'][i], 2)
            ws[col + '8'].border = f_border

            ws[col + '9'].font = name_font
            ws[col + '9'].alignment = c_c_alignment
            ws[col + '9'] = str(round(report['reporting_period']['increment_rates'][i] * 100, 2)) + "%" \
                if report['reporting_period']['increment_rates'][i] is not None else "-"
            ws[col + '9'].border = f_border

        # TCE TCO2E
        end_col = col
        # TCE
        tce_col = chr(ord(end_col) + 1)
        ws[tce_col + '7'].fill = table_fill
        ws[tce_col + '7'].font = name_font
        ws[tce_col + '7'].alignment = c_c_alignment
        ws[tce_col + '7'] = _('Ton of Standard Coal') + ' (TCE)'
        ws[tce_col + '7'].border = f_border

        ws[tce_col + '8'].font = name_font
        ws[tce_col + '8'].alignment = c_c_alignment
        ws[tce_col + '8'] = round(report['reporting_period']['total_in_kgce'] / 1000, 2) if \
            report['reporting_period']['increment_rate_in_kgce'] is not None else "-"
        ws[tce_col + '8'].border = f_border

        ws[tce_col + '9'].font = name_font
        ws[tce_col + '9'].alignment = c_c_alignment
        ws[tce_col + '9'] = str(round(report['reporting_period']['increment_rate_in_kgce'] * 100, 2)) + "%" \
            if report['reporting_period']['increment_rate_in_kgce'] is not None else "-"
        ws[tce_col + '9'].border = f_border

        # TCO2E
        tco2e_col = chr(ord(end_col) + 2)
        ws[tco2e_col + '7'].fill = table_fill
        ws[tco2e_col + '7'].font = name_font
        ws[tco2e_col + '7'].alignment = c_c_alignment
        ws[tco2e_col + '7'] = _('Ton of Carbon Dioxide Emissions') + ' (TCO2E)'
        ws[tco2e_col + '7'].border = f_border

        ws[tco2e_col + '8'].font = name_font
        ws[tco2e_col + '8'].alignment = c_c_alignment
        ws[tco2e_col + '8'] = round(report['reporting_period']['total_in_kgco2e'] / 1000, 2)if \
            report['reporting_period']['total_in_kgco2e'] is not None else "-"
        ws[tco2e_col + '8'].border = f_border

        ws[tco2e_col + '9'].font = name_font
        ws[tco2e_col + '9'].alignment = c_c_alignment
        ws[tco2e_col + '9'] = str(round(report['reporting_period']['increment_rate_in_kgco2e'] * 100, 2)) + "%" \
            if report['reporting_period']['increment_rate_in_kgco2e'] is not None else "-"
        ws[tco2e_col + '9'].border = f_border
    ####################################################################################################################
    # Second: Electricity Consumption by Time-Of-Use
    # 12: title
    # 13: table title
    # 14~17 table_data
    # Total: 6 rows
    ####################################################################################################################
    if "toppeaks" not in report['reporting_period'].keys() or \
            report['reporting_period']['toppeaks'] is None or \
            len(report['reporting_period']['toppeaks']) == 0:
        for i in range(12, 18 + 1):
            ws.row_dimensions[i].height = 0.1
    else:
        ws['B12'].font = title_font
        ws['B12'] = name + ' ' + _('Electricity Consumption by Time-Of-Use')

        ws.row_dimensions[13].height = 60
        ws['B13'].fill = table_fill
        ws['B13'].font = name_font
        ws['B13'].alignment = c_c_alignment
        ws['B13'].border = f_border

        ws['C13'].fill = table_fill
        ws['C13'].font = name_font
        ws['C13'].alignment = c_c_alignment
        ws['C13'].border = f_border
        ws['C13'] = _('Electricity Consumption by Time-Of-Use')

        ws['B14'].font = title_font
        ws['B14'].alignment = c_c_alignment
        ws['B14'] = _('TopPeak')
        ws['B14'].border = f_border

        ws['C14'].font = title_font
        ws['C14'].alignment = c_c_alignment
        ws['C14'].border = f_border
        ws['C14'] = round(report['reporting_period']['toppeaks'][0], 2)

        ws['B15'].font = title_font
        ws['B15'].alignment = c_c_alignment
        ws['B15'] = _('OnPeak')
        ws['B15'].border = f_border

        ws['C15'].font = title_font
        ws['C15'].alignment = c_c_alignment
        ws['C15'].border = f_border
        ws['C15'] = round(report['reporting_period']['onpeaks'][0], 2)

        ws['B16'].font = title_font
        ws['B16'].alignment = c_c_alignment
        ws['B16'] = _('MidPeak')
        ws['B16'].border = f_border

        ws['C16'].font = title_font
        ws['C16'].alignment = c_c_alignment
        ws['C16'].border = f_border
        ws['C16'] = round(report['reporting_period']['midpeaks'][0], 2)

        ws['B17'].font = title_font
        ws['B17'].alignment = c_c_alignment
        ws['B17'] = _('OffPeak')
        ws['B17'].border = f_border

        ws['C17'].font = title_font
        ws['C17'].alignment = c_c_alignment
        ws['C17'].border = f_border
        ws['C17'] = round(report['reporting_period']['offpeaks'][0], 2)

        pie = PieChart()
        pie.title = name + ' ' + _('Electricity Consumption by Time-Of-Use')
        labels = Reference(ws, min_col=2, min_row=14, max_row=17)
        pie_data = Reference(ws, min_col=3, min_row=13, max_row=17)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.25
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True

        ws.add_chart(pie, "D13")

    ####################################################################################################################
    # Third: Ton of Standard Coal(TCE) by Energy Category
    # current_row_number: title
    # current_row_number + 1: table title
    # current_row_number + 1 + ca_len table_data
    # Total: 2 + ca_len rows
    ####################################################################################################################
    current_row_number = 19
    if "subtotals_in_kgce" not in report['reporting_period'].keys() or \
            report['reporting_period']['subtotals_in_kgce'] is None or \
            len(report['reporting_period']['subtotals_in_kgce']) == 0:
        pass
    else:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Ton of Standard Coal(TCE) by Energy Category')

        current_row_number += 1
        table_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border

        ws['C' + str(current_row_number)].fill = table_fill
        ws['C' + str(current_row_number)].font = name_font
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)] = _('Ton of Standard Coal(TCE) by Energy Category')

        current_row_number += 1

        ca_len = len(report['reporting_period']['names'])

        for i in range(0, ca_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)] = report['reporting_period']['names'][i]
            ws['B' + str(current_row_number)].border = f_border

            ws['C' + str(current_row_number)].font = title_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(report['reporting_period']['subtotals_in_kgce'][i] / 1000, 3)

            current_row_number += 1

        table_end_row_number = current_row_number - 1

        pie = PieChart()
        pie.title = name + ' ' + ws.cell(column=3, row=table_start_row_number).value
        labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
        pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.25
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        table_cell = 'D' + str(table_start_row_number)
        ws.add_chart(pie, table_cell)

        if ca_len < 4:
            current_row_number = current_row_number - ca_len + 4

        current_row_number += 1

    ####################################################################################################################
    # Fourth: Ton of Carbon Dioxide Emissions(TCO2E) by Energy Category
    # current_row_number: title
    # current_row_number + 1: table title
    # current_row_number + 1 + ca_len table_data
    # Total: 2 + ca_len rows
    ####################################################################################################################
    if "subtotals_in_kgco2e" not in report['reporting_period'].keys() or \
            report['reporting_period']['subtotals_in_kgco2e'] is None or \
            len(report['reporting_period']['subtotals_in_kgco2e']) == 0:
        pass
    else:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Ton of Carbon Dioxide Emissions(TCO2E) by Energy Category')

        current_row_number += 1
        table_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 75
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border

        ws['C' + str(current_row_number)].fill = table_fill
        ws['C' + str(current_row_number)].font = name_font
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)] = _('Ton of Carbon Dioxide Emissions(TCO2E) by Energy Category')

        current_row_number += 1

        ca_len = len(report['reporting_period']['names'])

        for i in range(0, ca_len):
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)] = report['reporting_period']['names'][i]
            ws['B' + str(current_row_number)].border = f_border

            ws['C' + str(current_row_number)].font = title_font
            ws['C' + str(current_row_number)].alignment = c_c_alignment
            ws['C' + str(current_row_number)].border = f_border
            ws['C' + str(current_row_number)] = round(report['reporting_period']['subtotals_in_kgco2e'][i] / 1000, 3)
            current_row_number += 1

        table_end_row_number = current_row_number - 1

        pie = PieChart()
        pie.title = name + ' ' + ws.cell(column=3, row=table_start_row_number).value
        labels = Reference(ws, min_col=2, min_row=table_start_row_number + 1, max_row=table_end_row_number)
        pie_data = Reference(ws, min_col=3, min_row=table_start_row_number, max_row=table_end_row_number)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 7.75
        pie.width = 9
        s1 = pie.series[0]
        s1.dLbls = DataLabelList()
        s1.dLbls.showCatName = False
        s1.dLbls.showVal = True
        s1.dLbls.showPercent = True
        table_cell = 'D' + str(table_start_row_number)
        ws.add_chart(pie, table_cell)

        if ca_len < 4:
            current_row_number = current_row_number - ca_len + 4

        current_row_number += 1

    ####################################################################################################################
    # Fifth: Detailed Data
    # current_row_number: title
    # current_row_number+1 ~ current_row_number+1+ca_len*6-1: line
    # current_row_number+1+ca_len*6: table title
    # current_row_number+1+ca_len*6~: table_data
    ####################################################################################################################
    times = report['reporting_period']['timestamps']
    ca_len = len(report['reporting_period']['names'])
    real_timestamps_len = timestamps_data_not_equal_0(report['parameters']['timestamps'])
    table_row = current_row_number + 2 + ca_len * 6 + real_timestamps_len * 6
    chart_start_row_number = current_row_number + 1
    if "timestamps" not in report['reporting_period'].keys() or \
            report['reporting_period']['timestamps'] is None or \
            len(report['reporting_period']['timestamps']) == 0:
        pass
    else:
        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Detailed Data')

        ws.row_dimensions[table_row].height = 60
        ws['B'+str(table_row)].fill = table_fill
        ws['B' + str(table_row)].font = title_font
        ws['B'+str(table_row)].border = f_border
        ws['B'+str(table_row)].alignment = c_c_alignment
        ws['B'+str(table_row)] = _('Datetime')
        time = times[0]
        has_data = False
        max_row = 0
        if len(time) > 0:
            has_data = True
            max_row = table_row + len(time)

        if has_data:
            for i in range(0, len(time)):
                col = 'B'
                row = str(table_row+1 + i)
                # col = chr(ord('B') + i)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = time[i]
                ws[col + row].border = f_border

            for i in range(0, ca_len):
                # 38 title
                col = chr(ord('C') + i)

                ws[col + str(table_row)].fill = table_fill
                ws[col + str(table_row)].font = title_font
                ws[col + str(table_row)].alignment = c_c_alignment
                ws[col + str(table_row)] = report['reporting_period']['names'][i] + \
                    " (" + report['reporting_period']['units'][i] + ")"
                ws[col + str(table_row)].border = f_border

                # 39 data
                time = times[i]
                time_len = len(time)

                for j in range(0, time_len):
                    row = str(table_row+1 + j)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = round(report['reporting_period']['values'][i][j], 2)
                    ws[col + row].border = f_border

            current_row_number = table_row + 1 + len(times[0])

            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)].alignment = c_c_alignment
            ws['B' + str(current_row_number)].border = f_border
            ws['B' + str(current_row_number)] = _('Subtotal')

            for i in range(0, ca_len):
                col = chr(ord('C') + i)
                ws[col + str(current_row_number)].font = title_font
                ws[col + str(current_row_number)].alignment = c_c_alignment
                ws[col + str(current_row_number)].border = f_border
                ws[col + str(current_row_number)] = round(report['reporting_period']['subtotals'][i], 2)

                # line
                line = LineChart()
                line.title = _('Reporting Period Consumption') + ' - ' + ws.cell(column=3+i, row=table_row).value
                labels = Reference(ws, min_col=2, min_row=table_row+1, max_row=max_row)
                line_data = Reference(ws, min_col=3 + i, min_row=table_row, max_row=max_row)  # openpyxl bug
                line.add_data(line_data, titles_from_data=True)
                line.set_categories(labels)
                line_data = line.series[0]
                line_data.marker.symbol = "circle"
                line_data.smooth = True
                line.x_axis.crosses = 'min'
                line.height = 8.25
                line.width = 24
                line.dLbls = DataLabelList()
                line.dLbls.dLblPos = 't'
                line.dLbls.showVal = True  # val show
                line.dLbls.showPercent = True  # percent show
                chart_col = 'B'
                chart_cell = chart_col + str(chart_start_row_number + 6*i)
                ws.add_chart(line, chart_cell)

    ####################################################################################################################
    current_row_number += 2
    if "associated_equipment" not in report.keys() or \
            "energy_category_names" not in report['associated_equipment'].keys() or \
            len(report['associated_equipment']["energy_category_names"]) == 0 \
            or 'associated_equipment_names_array' not in report['associated_equipment'].keys() \
            or report['associated_equipment']['associated_equipment_names_array'] is None \
            or len(report['associated_equipment']['associated_equipment_names_array']) == 0 \
            or len(report['associated_equipment']['associated_equipment_names_array'][0]) == 0:
        pass
    else:
        associated_equipment = report['associated_equipment']

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)] = name + ' ' + _('Associated Equipment Data')

        current_row_number += 1
        table_start_row_number = current_row_number

        ws.row_dimensions[current_row_number].height = 60
        ws['B' + str(current_row_number)].fill = table_fill
        ws['B' + str(current_row_number)].font = name_font
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)] = _('Associated Equipment')
        ca_len = len(associated_equipment['energy_category_names'])

        for i in range(0, ca_len):
            row = chr(ord('C') + i)
            ws[row + str(current_row_number)].fill = table_fill
            ws[row + str(current_row_number)].font = name_font
            ws[row + str(current_row_number)].alignment = c_c_alignment
            ws[row + str(current_row_number)].border = f_border
            ws[row + str(current_row_number)] = \
                report['reporting_period']['names'][i] + " (" + report['reporting_period']['units'][i] + ")"

        associated_equipment_len = len(associated_equipment['associated_equipment_names_array'][0])

        for i in range(0, associated_equipment_len):
            current_row_number += 1
            row = str(current_row_number)

            ws['B' + row].font = title_font
            ws['B' + row].alignment = c_c_alignment
            ws['B' + row] = associated_equipment['associated_equipment_names_array'][0][i]
            ws['B' + row].border = f_border

            for j in range(0, ca_len):
                col = chr(ord('C') + j)
                ws[col + row].font = title_font
                ws[col + row].alignment = c_c_alignment
                ws[col + row] = round(associated_equipment['subtotals_array'][j][i], 2)
                ws[col + row].border = f_border
    ####################################################################################################################
    current_sheet_parameters_row_number = chart_start_row_number + ca_len * 6
    ####################################################################################################################
    # Sixth: Associated Equipment Detailed Data
    # current_row_number: title
    # current_row_number+1 ~ current_row_number+1+ca_len*6-1: line
    # current_row_number+1+ca_len*6: table title
    # current_row_number+1+ca_len*6~: table_data
    ####################################################################################################################

    associated_reporting_period_data_list = report['associated_report_period_list']
    for associated_reporting_period_data in associated_reporting_period_data_list:
        current_row_number = current_row_number + 1 
        times = associated_reporting_period_data['timestamps']
        ca_len = len(associated_reporting_period_data['names'])
        table_row = current_row_number + 2 + ca_len * 6
        chart_start_row_number = current_row_number + 1

        if "timestamps" not in associated_reporting_period_data.keys() or \
                associated_reporting_period_data['timestamps'] is None or \
                len(associated_reporting_period_data['timestamps']) == 0:
            pass
        else:
            ws['B' + str(current_row_number)].font = title_font
            ws['B' + str(current_row_number)] = \
                str(report['associated_equipment']['associated_equipment_names_array'][0]
                    [associated_reporting_period_data_list.index(associated_reporting_period_data)]) + \
                ' ' + _('Detailed Data')

            ws.row_dimensions[table_row].height = 60
            ws['B'+str(table_row)].fill = table_fill
            ws['B' + str(table_row)].font = title_font
            ws['B'+str(table_row)].border = f_border
            ws['B'+str(table_row)].alignment = c_c_alignment
            ws['B'+str(table_row)] = _('Datetime')
            time = times[0]
            has_data = False
            max_row = 0
            if len(time) > 0:
                has_data = True
                max_row = table_row + len(time)

            if has_data:
                for i in range(0, len(time)):
                    col = 'B'
                    row = str(table_row+1 + i)
                    # col = chr(ord('B') + i)
                    ws[col + row].font = title_font
                    ws[col + row].alignment = c_c_alignment
                    ws[col + row] = time[i]
                    ws[col + row].border = f_border

                for i in range(0, ca_len):
                    col = chr(ord('C') + i)

                    ws[col + str(table_row)].fill = table_fill
                    ws[col + str(table_row)].font = title_font
                    ws[col + str(table_row)].alignment = c_c_alignment
                    ws[col + str(table_row)] = associated_reporting_period_data['names'][i] + \
                        " (" + associated_reporting_period_data['units'][i] + ")"
                    ws[col + str(table_row)].border = f_border

                    time = times[i]
                    time_len = len(time)

                    for j in range(0, time_len):
                        row = str(table_row+1 + j)

                        ws[col + row].font = title_font
                        ws[col + row].alignment = c_c_alignment
                        ws[col + row] = round(associated_reporting_period_data['values'][i][j], 2)
                        ws[col + row].border = f_border

                current_row_number = table_row + 1 + len(times[0])

                ws['B' + str(current_row_number)].font = title_font
                ws['B' + str(current_row_number)].alignment = c_c_alignment
                ws['B' + str(current_row_number)].border = f_border
                ws['B' + str(current_row_number)] = _('Subtotal')

                for i in range(0, ca_len):
                    col = chr(ord('C') + i)
                    ws[col + str(current_row_number)].font = title_font
                    ws[col + str(current_row_number)].alignment = c_c_alignment
                    ws[col + str(current_row_number)].border = f_border
                    ws[col + str(current_row_number)] = round(associated_reporting_period_data['subtotals'][i], 2)

                    # line
                    line = LineChart()
                    line.title = _('Reporting Period Consumption') + ' - ' + ws.cell(column=3+i, row=table_row).value
                    labels = Reference(ws, min_col=2, min_row=table_row+1, max_row=max_row)
                    line_data = Reference(ws, min_col=3 + i, min_row=table_row, max_row=max_row)  # openpyxl bug
                    line.add_data(line_data, titles_from_data=True)
                    line.set_categories(labels)
                    line_data = line.series[0]
                    line_data.marker.symbol = "circle"
                    line_data.smooth = True
                    line.x_axis.crosses = 'min'
                    line.height = 8.25
                    line.width = 24
                    line.dLbls = DataLabelList()
                    line.dLbls.dLblPos = 't'
                    line.dLbls.showVal = True  # val show
                    line.dLbls.showPercent = True  # percent show
                    chart_col = 'B'
                    chart_cell = chart_col + str(chart_start_row_number + 6*i)
                    ws.add_chart(line, chart_cell)

    ####################################################################################################################
    if 'parameters' not in report.keys() or \
            report['parameters'] is None or \
            'names' not in report['parameters'].keys() or \
            report['parameters']['names'] is None or \
            len(report['parameters']['names']) == 0 or \
            'timestamps' not in report['parameters'].keys() or \
            report['parameters']['timestamps'] is None or \
            len(report['parameters']['timestamps']) == 0 or \
            'values' not in report['parameters'].keys() or \
            report['parameters']['values'] is None or \
            len(report['parameters']['values']) == 0 or \
            timestamps_data_all_equal_0(report['parameters']['timestamps']):
        pass
    else:
        ################################################################################################################
        # new worksheet
        ################################################################################################################

        parameters_data = report['parameters']
        parameters_names_len = len(parameters_data['names'])

        file_name = (re.sub(r'[^A-Z]', '', ws.title))+'_'
        parameters_ws = wb.create_sheet(file_name + _('Parameters'))

        parameters_timestamps_data_max_len = \
            get_parameters_timestamps_lists_max_len(list(parameters_data['timestamps']))

        # Row height
        parameters_ws.row_dimensions[1].height = 102
        for i in range(2, 7 + 1):
            parameters_ws.row_dimensions[i].height = 42

        for i in range(8, parameters_timestamps_data_max_len + 10):
            parameters_ws.row_dimensions[i].height = 60

        # Col width
        parameters_ws.column_dimensions['A'].width = 1.5

        parameters_ws.column_dimensions['B'].width = 25.0

        for i in range(3, 12+parameters_names_len*3):
            parameters_ws.column_dimensions[format_cell.get_column_letter(i)].width = 15.0

        # Img
        img = Image("excelexporters/myems.png")
        parameters_ws.add_image(img, 'A1')

        # Title
        parameters_ws['B3'].alignment = b_r_alignment
        parameters_ws['B3'] = _('Name') + ':'
        parameters_ws['C3'].border = b_border
        parameters_ws['C3'].alignment = b_c_alignment
        parameters_ws['C3'] = name

        parameters_ws['D3'].alignment = b_r_alignment
        parameters_ws['D3'] = _('Period Type') + ':'
        parameters_ws['E3'].border = b_border
        parameters_ws['E3'].alignment = b_c_alignment
        parameters_ws['E3'] = period_type

        parameters_ws['B4'].alignment = b_r_alignment
        parameters_ws['B4'] = _('Reporting Start Datetime') + ':'
        parameters_ws['C4'].border = b_border
        parameters_ws['C4'].alignment = b_c_alignment
        parameters_ws['C4'] = reporting_start_datetime_local

        parameters_ws['D4'].alignment = b_r_alignment
        parameters_ws['D4'] = _('Reporting End Datetime') + ':'
        parameters_ws['E4'].border = b_border
        parameters_ws['E4'].alignment = b_c_alignment
        parameters_ws['E4'] = reporting_end_datetime_local

        parameters_ws_current_row_number = 6

        parameters_ws['B' + str(parameters_ws_current_row_number)].font = title_font
        parameters_ws['B' + str(parameters_ws_current_row_number)] = name + ' ' + _('Parameters')

        parameters_ws_current_row_number += 1

        parameters_table_start_row_number = parameters_ws_current_row_number

        parameters_ws.row_dimensions[parameters_ws_current_row_number].height = 80

        parameters_ws_current_row_number += 1

        table_current_col_number = 2

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue
            col = format_cell.get_column_letter(table_current_col_number)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border

            col = format_cell.get_column_letter(table_current_col_number + 1)

            parameters_ws[col + str(parameters_ws_current_row_number-1)].fill = table_fill
            parameters_ws[col + str(parameters_ws_current_row_number-1)].border = f_border
            parameters_ws[col + str(parameters_ws_current_row_number-1)].font = name_font
            parameters_ws[col + str(parameters_ws_current_row_number-1)].alignment = c_c_alignment
            parameters_ws[col + str(parameters_ws_current_row_number-1)] = parameters_data['names'][i]

            table_current_row_number = parameters_ws_current_row_number

            for j, value in enumerate(list(parameters_data['timestamps'][i])):
                col = format_cell.get_column_letter(table_current_col_number)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = value

                col = format_cell.get_column_letter(table_current_col_number + 1)

                parameters_ws[col + str(table_current_row_number)].border = f_border
                parameters_ws[col + str(table_current_row_number)].font = title_font
                parameters_ws[col + str(table_current_row_number)].alignment = c_c_alignment
                parameters_ws[col + str(table_current_row_number)] = round(parameters_data['values'][i][j], 2)

                table_current_row_number += 1

            table_current_col_number = table_current_col_number + 3

        ########################################################
        # parameters chart and parameters table
        ########################################################

        ws['B' + str(current_sheet_parameters_row_number)].font = title_font
        ws['B' + str(current_sheet_parameters_row_number)] = name + ' ' + _('Parameters')

        current_sheet_parameters_row_number += 1

        chart_start_row_number = current_sheet_parameters_row_number

        col_index = 0

        for i in range(0, parameters_names_len):

            if len(parameters_data['timestamps'][i]) == 0:
                continue

            line = LineChart()
            data_col = 3+col_index*3
            labels_col = 2+col_index*3
            col_index += 1
            line.title = _('Parameters') + ' - ' + \
                parameters_ws.cell(row=parameters_table_start_row_number, column=data_col).value
            labels = Reference(parameters_ws, min_col=labels_col, min_row=parameters_table_start_row_number + 1,
                               max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line_data = Reference(parameters_ws, min_col=data_col, min_row=parameters_table_start_row_number,
                                  max_row=(len(parameters_data['timestamps'][i])+parameters_table_start_row_number))
            line.add_data(line_data, titles_from_data=True)
            line.set_categories(labels)
            line_data = line.series[0]
            line_data.marker.symbol = "circle"
            line_data.smooth = True
            line.x_axis.crosses = 'min'
            line.height = 8.25
            line.width = 24
            line.dLbls = DataLabelList()
            line.dLbls.dLblPos = 't'
            line.dLbls.showVal = False
            line.dLbls.showPercent = False
            chart_col = 'B'
            chart_cell = chart_col + str(chart_start_row_number)
            chart_start_row_number += 6
            ws.add_chart(line, chart_cell)

        current_sheet_parameters_row_number = chart_start_row_number

        current_sheet_parameters_row_number += 1
    ####################################################################################################################
    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename


def get_parameters_timestamps_lists_max_len(parameters_timestamps_lists):
    max_len = 0
    for i, value in enumerate(list(parameters_timestamps_lists)):
        if len(value) > max_len:
            max_len = len(value)

    return max_len


def timestamps_data_all_equal_0(lists):
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            return False

    return True


def timestamps_data_not_equal_0(lists):
    number = 0
    for i, value in enumerate(list(lists)):
        if len(value) > 0:
            number += 1
    return number

